"""Allocate Deployment Tracker — main entry point."""

from __future__ import annotations
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl
import yaml
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
from rules import classify, normalise_schedule  # noqa: E402
from calendar_logic import compute_tranche_dates, load_holidays, to_date  # noqa: E402

ROOT = HERE.parent
CONFIG_PATH = ROOT / "config.yaml"
HOLIDAYS_PATH = ROOT / "data" / "holidays.json"
STATE_PATH = ROOT / "state" / "last_run.json"
OUTPUTS_DIR = ROOT / "outputs"


def load_config() -> dict:
    with open(CONFIG_PATH) as f:
        cfg = yaml.safe_load(f)
    cfg["_models_in_scope_lower"] = {m.strip().lower() for m in cfg["models_in_scope"]}
    for r in cfg["asset_class_rules"]:
        r["schedule"] = normalise_schedule(r.get("schedule"))
    return cfg


def load_model_master(model_master_path: Path, models_in_scope: set) -> dict:
    print(f"Loading Model Master ({model_master_path.name})...", flush=True)
    wb = openpyxl.load_workbook(model_master_path, data_only=True, read_only=True)
    cache = {}
    for sn in wb.sheetnames:
        if sn.strip().lower() not in models_in_scope:
            continue
        ws = wb[sn]
        secs = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            if not any(row[j] for j in range(min(len(row), 9))):
                continue
            tw = row[8] if len(row) > 8 else 0
            secs.append({
                "asset_class": row[1], "category": row[2], "instrument": row[3],
                "name": row[5], "isin": row[6], "target_weight": tw or 0,
            })
        cache[sn.strip().lower()] = secs
        print(f"  Cached '{sn}': {len(secs)} securities", flush=True)
    wb.close()
    return cache


def load_client_events(client_repo_path: Path) -> list:
    print(f"\nLoading Client Repository ({client_repo_path.name})...", flush=True)
    wb = openpyxl.load_workbook(client_repo_path, data_only=True, read_only=True)
    ws = wb["Repository"]
    events = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row[2]:
            continue
        events.append({
            "row_idx": idx + 2,  # 1-based excel row number for stable identification
            "custody_code": row[0], "client_code": row[1], "client_name": row[2],
            "scheme": row[3], "fee_model": row[4], "trans_type": row[5],
            "amount": row[6] or 0, "activation_date": row[7], "stp": row[8],
            "ws_client_id": row[9], "date_deployed": row[10],
            "model_assigned": row[11], "model_id": row[12], "rm": row[13],
            "rm_region": row[14], "team_lead": row[15],
            "division": row[16] if len(row) > 16 else None,
        })
    wb.close()
    print(f"  Loaded {len(events)} events", flush=True)
    return events


def load_state() -> dict:
    if STATE_PATH.exists():
        with open(STATE_PATH) as f:
            return json.load(f)
    return {"completed_tranches": {}}


def save_state(state: dict) -> None:
    STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(STATE_PATH, "w") as f:
        json.dump(state, f, indent=2, default=str)


def event_id(e: dict) -> str:
    """Stable unique identifier per deployment event.

    Built from a composite of fields so every row gets a unique ID even when:
      - ws_client_id is None (268+ rows in the Client Repo have no WS ID), or
      - the same WS_ID is shared by two distinct clients with the same Date Deployed.
    The composite is deterministic from row data, so re-runs assign the same ID to the
    same row, preserving checkbox state in the browser.
    """
    dd = to_date(e["date_deployed"]).isoformat()
    ws = e.get("ws_client_id") or "NW"
    cust = (e.get("custody_code") or "").replace("|", "_")
    cc = (e.get("client_code") or "").replace("|", "_")
    amt = int(e.get("amount") or 0)
    idx = e.get("row_idx", 0)
    return f"{ws}|{dd}|{cust}|{cc}|{amt}|r{idx}"


def tranche_key(security_isin, day_offset: int) -> str:
    return f"{security_isin}@D{day_offset}"


def compute_event_tranches(event, model_secs, rules, holidays) -> dict:
    day0 = to_date(event["date_deployed"])
    corpus = float(event["amount"])
    tranches = []
    parking = defaultdict(float)
    cash_residual = 0.0
    liquid_end_state = defaultdict(float)

    for s in model_secs:
        tw = s["target_weight"] or 0
        target_inr = corpus * tw / 100.0
        rule = classify(rules, s["asset_class"], s["instrument"], s["category"], s["name"], event_date=day0)
        bucket = rule["name"]
        schedule = rule["schedule"]
        parking_fund = rule.get("parking_fund")

        if bucket == "Cash (residual)":
            if tw > 0:
                cash_residual += target_inr
            continue
        if bucket == "Liquid ETF (parking)":
            if tw > 0:
                liquid_end_state[s["name"]] += target_inr
            continue
        if tw == 0:
            continue
        if not schedule:
            tranches.append({
                "bucket": bucket, "security": s["name"], "isin": s["isin"],
                "target_weight": tw, "target_inr": round(target_inr, 2),
                "day_offset": None, "tranche_pct": None,
                "tranche_inr": round(target_inr, 2),
                "redemption_date": None, "deployment_date": None,
                "parking_fund": parking_fund,
                "tranche_key": tranche_key(s["isin"] or s["name"], -1),
            })
            continue

        for day_offset, pct in schedule.items():
            tranche_inr = target_inr * pct
            redem_dt, deploy_dt = compute_tranche_dates(day0, day_offset, holidays)
            tranches.append({
                "bucket": bucket, "security": s["name"], "isin": s["isin"],
                "target_weight": tw, "target_inr": round(target_inr, 2),
                "day_offset": day_offset, "tranche_pct": pct,
                "tranche_inr": round(tranche_inr, 2),
                "redemption_date": redem_dt, "deployment_date": deploy_dt,
                "parking_fund": parking_fund,
                "tranche_key": tranche_key(s["isin"] or s["name"], day_offset),
            })
            if day_offset == 0 and parking_fund:
                day0_pct = sum(p for d, p in schedule.items() if d == 0)
                undeployed = 1 - day0_pct
                if undeployed > 0:
                    parking[parking_fund] += target_inr * undeployed

    return {
        "tranches": tranches, "parking": dict(parking),
        "cash_residual": round(cash_residual, 2),
        "liquid_end_state": dict(liquid_end_state),
    }


def status_for_tranche(t, today, completed_set):
    """Status rules:
    - Tranche key in completed_set OR deploy_date strictly before today  → Done
    - deploy_date is None → —
    - deploy_date in current week (Mon-Sun) → Due this week
    - else → Upcoming
    Assumption: if the deployment date has passed, the deployment has been completed.
    """
    deploy_dt = t["deployment_date"]
    if t["tranche_key"] in completed_set:
        return "Done"
    if deploy_dt is None:
        return "—"
    if deploy_dt < today:
        return "Done"
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    if week_start <= deploy_dt <= week_end:
        return "Due this week"
    return "Upcoming"


def _serialize(obj):
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    raise TypeError(f"Not serializable: {type(obj)}")


def write_dataset_json(in_scope, out_of_scope, today):
    print("\nWriting dataset.json...", flush=True)
    data = {
        "generated_at": datetime.now().isoformat(),
        "today": today.isoformat(),
        "events": [], "out_of_scope": [],
    }
    for item in in_scope:
        e, r = item["event"], item["result"]
        data["events"].append({
            "event_id": item["event_id"],
            "ws_client_id": e["ws_client_id"],
            "client_code": e.get("client_code"),
            "custody_code": e.get("custody_code"),
            "client_name": e["client_name"],
            "model": e["model_assigned"], "scheme": e["scheme"],
            "trans_type": e["trans_type"], "rm": e["rm"],
            "team_lead": e["team_lead"], "division": e["division"],
            "corpus": e["amount"],
            "date_deployed": to_date(e["date_deployed"]).isoformat(),
            "pre_cutoff": item.get("pre_cutoff", False),
            "parking": r["parking"], "cash_residual": r["cash_residual"],
            "liquid_end_state": r["liquid_end_state"],
            "tranches": [{
                **t,
                "redemption_date": t["redemption_date"].isoformat() if t["redemption_date"] else None,
                "deployment_date": t["deployment_date"].isoformat() if t["deployment_date"] else None,
            } for t in r["tranches"]],
        })
    for o in out_of_scope:
        e = o["event"]
        data["out_of_scope"].append({
            "ws_client_id": e["ws_client_id"], "client_name": e["client_name"],
            "model": e.get("model_assigned"), "amount": e.get("amount"),
            "date_deployed": to_date(e["date_deployed"]).isoformat()
                if isinstance(e.get("date_deployed"), (date, datetime)) else None,
            "reason": o["reason"],
        })

    final_path = OUTPUTS_DIR / "dataset.json"
    import tempfile, shutil, os
    fd, tp = tempfile.mkstemp(suffix=".json"); os.close(fd)
    with open(tp, "w") as f:
        json.dump(data, f, separators=(",", ":"), default=_serialize)
    shutil.copy(tp, final_path)
    Path(tp).unlink(missing_ok=True)
    print(f"  Wrote {final_path} ({final_path.stat().st_size:,} bytes)", flush=True)


def write_excel_tracker(in_scope, out_of_scope, today):
    print("\nWriting Excel tracker...", flush=True)
    final_path = OUTPUTS_DIR / f"Deployment_Tracker_{today.strftime('%Y%m%d')}.xlsx"

    HDR = Font(bold=True, color="FFFFFF", size=11)
    HDR_FILL = PatternFill("solid", fgColor="1F4E79")
    MONEY = "#,##0"

    import tempfile, shutil, os
    fd, tp = tempfile.mkstemp(suffix=".xlsx"); os.close(fd)
    out_path = Path(tp)

    wb = openpyxl.Workbook()

    def style_header(ws, row_idx):
        for c in ws[row_idx]:
            c.font = HDR; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center")

    # Sheet 1: Status Rollup
    ws = wb.active; ws.title = "Status Rollup"
    ws.append(["WS ID", "Client", "Model", "Trans", "RM", "Corpus", "Date Deployed",
               "Day 0 Deploy", "Zerodha Park", "Kotak Park", "DSP Park",
               "% Deployed", "Tranches Done", "Tranches Total", "Next Action"])
    style_header(ws, 1)
    for item in in_scope:
        e, r = item["event"], item["result"]
        day0_dep = sum(t["tranche_inr"] for t in r["tranches"] if t["day_offset"] == 0)
        total_target = sum(t["tranche_inr"] for t in r["tranches"])
        deployed_so_far = sum(t["tranche_inr"] for t in r["tranches"] if t.get("status") == "Done")
        pct = (deployed_so_far / total_target) if total_target else 0
        done = sum(1 for t in r["tranches"] if t.get("status") == "Done")
        total = len(r["tranches"])
        nexts = sorted([t["deployment_date"] for t in r["tranches"]
                        if t.get("status") in ("Due this week", "Upcoming", "Overdue") and t["deployment_date"]])
        next_action = nexts[0].isoformat() if nexts else "—"
        ws.append([
            e["ws_client_id"], e["client_name"], e["model_assigned"], e["trans_type"], e["rm"],
            e["amount"], to_date(e["date_deployed"]),
            round(day0_dep), round(r["parking"].get("Zerodha", 0)),
            round(r["parking"].get("Kotak", 0)), round(r["parking"].get("DSP", 0)),
            round(pct * 100, 1), done, total, next_action,
        ])
    for col, w in zip("ABCDEFGHIJKLMNO", [10, 28, 20, 16, 18, 14, 13, 13, 13, 13, 13, 11, 9, 9, 13]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"

    # Sheet 2: All Tranches
    ws = wb.create_sheet("All Tranches")
    ws.append(["WS ID", "Client", "Model", "Date Deployed", "Bucket", "Security",
               "Target Wt %", "Target ₹", "Day", "Tranche %", "Tranche ₹",
               "Redem (Mon)", "Deploy (Thu)", "Parking Fund", "Status"])
    style_header(ws, 1)
    print("  Writing tranche rows...", flush=True)
    n_rows = 0
    for item in in_scope:
        e, r = item["event"], item["result"]
        for t in r["tranches"]:
            ws.append([
                e["ws_client_id"], e["client_name"], e["model_assigned"],
                to_date(e["date_deployed"]),
                t["bucket"], t["security"], t["target_weight"], t["target_inr"],
                f"Day {t['day_offset']}" if t["day_offset"] is not None else "—",
                t["tranche_pct"] if t["tranche_pct"] is not None else "",
                t["tranche_inr"], t["redemption_date"], t["deployment_date"],
                t["parking_fund"] or "", t.get("status", "—"),
            ])
            n_rows += 1
    print(f"  Appended {n_rows} tranche rows", flush=True)
    for col, w in zip("ABCDEFGHIJKLMNO", [10, 26, 20, 13, 22, 38, 10, 14, 8, 9, 14, 13, 13, 12, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "B2"

    # Sheet 3: This Week
    ws = wb.create_sheet("This Week")
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    ws["A1"] = f"Action Calendar: {week_start} to {week_end}"
    ws["A1"].font = Font(bold=True, size=12)
    ws.append([])
    ws.append(["Date", "Day", "Action", "Client", "Model", "Bucket / Security",
               "Amount (₹)", "Liquid Fund", "Status"])
    style_header(ws, 3)
    rows = []
    for item in in_scope:
        e, r = item["event"], item["result"]
        for t in r["tranches"]:
            for action, dt in [("Redeem", t["redemption_date"]), ("Deploy", t["deployment_date"])]:
                if dt and week_start <= dt <= week_end:
                    rows.append((dt, action, e, t))
    rows.sort(key=lambda x: (x[0], 0 if x[1] == "Redeem" else 1, x[2]["client_name"]))
    for dt, action, e, t in rows:
        wd = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"][dt.weekday()]
        ws.append([dt, wd, action, e["client_name"], e["model_assigned"],
                   f"{t['bucket']}: {t['security']}",
                   t["tranche_inr"], t["parking_fund"] or "", t.get("status", "—")])
    for col, w in zip("ABCDEFGHI", [12, 6, 8, 26, 20, 40, 14, 12, 14]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A4"

    # Sheet 4: Pending
    ws = wb.create_sheet("Pending")
    ws.append(["Days Overdue", "Status", "Client", "Model", "Bucket / Security",
               "Amount (₹)", "Redem (Mon)", "Deploy (Thu)", "Liquid Fund"])
    style_header(ws, 1)
    pending = []
    for item in in_scope:
        e, r = item["event"], item["result"]
        for t in r["tranches"]:
            st = t.get("status", "—")
            if st in ("Overdue", "Due this week"):
                days = (today - t["deployment_date"]).days if t["deployment_date"] and st == "Overdue" else 0
                pending.append({"days": days, "event": e, "tranche": t})
    pending.sort(key=lambda r: (-r["days"], r["event"]["client_name"]))
    for row in pending:
        e, t = row["event"], row["tranche"]
        ws.append([
            row["days"] if t.get("status") == "Overdue" else "—",
            t.get("status"), e["client_name"], e["model_assigned"],
            f"{t['bucket']}: {t['security']}",
            t["tranche_inr"], t["redemption_date"], t["deployment_date"],
            t["parking_fund"] or "",
        ])
    for col, w in zip("ABCDEFGHI", [13, 14, 26, 20, 40, 14, 13, 13, 12]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    # Sheet 5: Out of Scope
    ws = wb.create_sheet("Out of Scope")
    ws.append(["WS ID", "Client", "Model", "Amount", "Date Deployed", "Reason"])
    style_header(ws, 1)
    for o in out_of_scope:
        e = o["event"]
        ws.append([
            e.get("ws_client_id"), e.get("client_name"), e.get("model_assigned"),
            e.get("amount"),
            to_date(e["date_deployed"]) if isinstance(e.get("date_deployed"), (date, datetime)) else None,
            o["reason"],
        ])
    for col, w in zip("ABCDEF", [10, 26, 22, 14, 13, 30]):
        ws.column_dimensions[col].width = w

    print("  Saving workbook...", flush=True)
    wb.save(out_path)
    shutil.copy(out_path, final_path)
    out_path.unlink(missing_ok=True)
    print(f"  Wrote {final_path} ({final_path.stat().st_size:,} bytes)", flush=True)


def render_dashboard_step(today):
    print("\nRendering HTML dashboard...", flush=True)
    sys.path.insert(0, str(HERE))
    import render_dashboard
    render_dashboard.render(OUTPUTS_DIR / "dataset.json", OUTPUTS_DIR / "dashboard.html", today)


def main():
    today = date.today()
    OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
    cfg = load_config()
    holidays = load_holidays(HOLIDAYS_PATH)
    rules = cfg["asset_class_rules"]
    cutoff_str = cfg.get("tracking_cutoff_date")
    cutoff_date = date.fromisoformat(cutoff_str) if cutoff_str else date(1970, 1, 1)
    print(f"Active tracking cutoff: {cutoff_date}", flush=True)

    # Resolve paths: try configured path + common locations, and validate the file is a real xlsx
    # (not a OneDrive cloud-only placeholder).
    import zipfile
    candidates = [Path(cfg["allocate_folder"]), ROOT, ROOT.parent,
                  Path("/tmp"),
                  Path("/sessions/sleepy-keen-rubin/mnt/uploads")]

    def _is_valid_xlsx(p):
        try:
            with zipfile.ZipFile(p):
                return True
        except Exception:
            return False

    def _find(filename, fallback_globs=()):
        for cand in candidates:
            p = cand / filename
            if p.exists() and _is_valid_xlsx(p):
                return p
        for cand in candidates:
            if not cand.exists():
                continue
            for pattern in fallback_globs:
                for p in cand.glob(pattern):
                    if _is_valid_xlsx(p):
                        return p
        return None

    client_repo = _find(cfg["client_repo_filename"], ["03 Client Repository IPS v2*.xlsx"])
    model_master = _find(cfg["model_master_filename"], ["02 Model_Master_Allocate IPS*.xlsx"])
    if client_repo is None or model_master is None:
        print(f"ERROR: Source files not found or invalid.")
        print(f"  client_repo: {client_repo}")
        print(f"  model_master: {model_master}")
        print(f"  Searched: {[str(c) for c in candidates]}")
        sys.exit(1)
    print(f"Source files:")
    print(f"  Client Repo: {client_repo}")
    print(f"  Model Master: {model_master}")

    models_cache = load_model_master(model_master, cfg["_models_in_scope_lower"])
    events = load_client_events(client_repo)

    state = load_state()
    completed = state.get("completed_tranches", {})

    # ─── Identify Liquid DPMS clients (entire client excluded if any row matches) ───
    liquid_filter = cfg.get("liquid_dpms_filters", {}) or {}
    scheme_keywords = [k.lower() for k in liquid_filter.get("scheme_contains_all", [])]
    model_equals = (liquid_filter.get("model_equals") or "").strip().lower()
    liquid_client_names = set()
    for e in events:
        scheme = (e.get("scheme") or "").lower()
        model_lc = (e.get("model_assigned") or "").strip().lower()
        scheme_match = scheme_keywords and all(k in scheme for k in scheme_keywords)
        model_match = bool(model_equals) and model_lc == model_equals
        if scheme_match or model_match:
            if e.get("client_name"):
                liquid_client_names.add(e["client_name"])
    print(f"Liquid DPMS clients excluded by name: {len(liquid_client_names)}", flush=True)

    min_inflow = float(cfg.get("min_inflow_amount") or 0)
    print(f"Minimum inflow amount: \u20b9{min_inflow:,.0f}", flush=True)

    in_scope, out_of_scope = [], []
    print(f"\nProcessing {len(events)} events...", flush=True)
    for e in events:
        if e.get("client_name") in liquid_client_names:
            out_of_scope.append({"event": e, "reason": "client has Liquid DPMS entry"})
            continue
        if (e.get("amount") or 0) < min_inflow:
            out_of_scope.append({"event": e, "reason": f"amount < min_inflow (\u20b9{min_inflow:,.0f})"})
            continue
        if not e.get("model_assigned"):
            out_of_scope.append({"event": e, "reason": "no model assigned"})
            continue
        ma = e["model_assigned"].strip().lower()
        if ma not in cfg["_models_in_scope_lower"]:
            out_of_scope.append({"event": e, "reason": f"model '{e['model_assigned']}' out of scope"})
            continue
        if ma not in models_cache:
            out_of_scope.append({"event": e, "reason": f"model '{e['model_assigned']}' has no sheet"})
            continue
        if not isinstance(e["date_deployed"], (date, datetime)):
            out_of_scope.append({"event": e, "reason": "no Date Deployed"})
            continue

        result = compute_event_tranches(e, models_cache[ma], rules, holidays)
        eid = event_id(e)
        event_date = to_date(e["date_deployed"])
        pre_cutoff = event_date < cutoff_date

        if pre_cutoff:
            # Pre-cutoff: assume the entire deployment was completed historically.
            ev_completed = {t["tranche_key"] for t in result["tranches"]}
            completed[eid] = {k: True for k in ev_completed}
        elif eid in completed:
            ev_completed = set(completed.get(eid, {}).keys())
        else:
            # First-run active event: auto-mark tranches whose deploy date is past.
            ev_completed = {t["tranche_key"] for t in result["tranches"]
                            if t["deployment_date"] and t["deployment_date"] < today}
            if ev_completed:
                completed[eid] = {k: True for k in ev_completed}

        for t in result["tranches"]:
            t["status"] = status_for_tranche(t, today, ev_completed)
        in_scope.append({"event": e, "result": result, "event_id": eid, "pre_cutoff": pre_cutoff})

    print(f"In scope: {len(in_scope)} | Out of scope: {len(out_of_scope)}", flush=True)

    state["completed_tranches"] = completed
    state["last_run"] = today.isoformat()
    save_state(state)

    write_dataset_json(in_scope, out_of_scope, today)
    write_excel_tracker(in_scope, out_of_scope, today)
    render_dashboard_step(today)
    print(f"\n✓ Done. Outputs in {OUTPUTS_DIR}")


if __name__ == "__main__":
    main()
